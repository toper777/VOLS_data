#  Copyright (c) 2022. Tikhon Ostapenko
import configparser
import datetime
import json
import ssl
import sys
import urllib.request
from io import BytesIO
from pathlib import Path
from typing import List

import pandas as pd
import requests
from loguru import logger
from openpyxl.utils import get_column_letter
from pandas import DataFrame
from redmail import EmailSender

from Colors import Colors as Color
from FormattedWorkbook import FormattedWorkbook
from gdc_vols import PROGRAM_NAME, PROGRAM_VERSION

config_file = 'gdc_vols.ini'


def email_split(mail_list: str) -> list:
    """Возвращает список email адресов из строки"""
    email_list = mail_list.strip().split(',| |;')
    for i in range(len(email_list)):
        email_list[i] = email_list[i].strip()
    return email_list


def fill_cell_names():
    """
    Заполнение словаря для обращения к ячейкам Excel 1:A, 2:B,... 27:AA, 28:AB и так далее до ZZZ

    :return Dictionary:
    """
    _count = 1
    _cell_names = {}

    for _i in range(65, 91):
        _cell_names[_count] = chr(_i)
        _count += 1
    for _i in range(65, 91):
        for _j in range(65, 91):
            _cell_names[_count] = chr(_i) + chr(_j)
            _count += 1
    for _i in range(65, 91):
        for _j in range(65, 91):
            for _k in range(65, 91):
                _cell_names[_count] = chr(_i) + chr(_j) + chr(_k)
                _count += 1
    return _cell_names


def print_debug(level, message):
    """
    Функция печати DEBUG сообщений.
    Принимает параметр номер и само сообщение.
    Сообщение должно быть в формате строки вывода

    :param level:
    :param message:
    """
    print(f'{Color.RED}DEBUG ({level}): \n{Color.END}{Color.YELLOW}{message}{Color.END}')


def read_from_dashboard(_url: str, data_type: str = "JSON", check_ssl: bool = True) -> pd.DataFrame:
    """
    Читает данные JSON или Excel из url и сохраняет их в DataFrame

    :param _url: Местоположения данных:
    :param data_type: Тип получаемых данных 'JSON' млм 'EXCEL':
    :param check_ssl: Проверка сертификата (True или False):
    :return DataFrame:
    """
    print(f'Получаем данные из: "{_url}"')
    try:
        # Временно выключаем проверку сертификатов
        # ssl._create_default_https_context = ssl._create_unverified_context
        # Временно выключаем проверку сертификатов

        if data_type.lower() == "excel":
            _dashboard_data = pd.read_excel(_url, parse_dates=True)
        else:
            # Временно выключаем проверку сертификатов
            response = requests.get(_url, verify=check_ssl)
            _dashboard_data = pd.DataFrame(response.json())
            # Временно выключаем проверку сертификатов

            # _dashboard_data = pd.read_json(_url, convert_dates=['дата', 'Дата'])
    except Exception as e:
        print(f"ERROR: can't read data from url {_url}. {e}")
        sys.exit(1)
    return _dashboard_data


def get_update_date(_url, check_ssl=True):
    """Читает дату обновления через API из url и возвращает ее."""

    print(f'Получаем дату обновления данных из: "{_url}"')
    try:
        # Временно выключаем проверку сертификатов
        response = requests.get(_url, verify=check_ssl)
        data_json = json.loads(response.content.decode('utf-8'))
        # Временно выключаем проверку сертификатов

        # response = urllib.request.urlopen(_url)
        # data_json = json.load(response)
    except Exception as e:
        print(f"ERROR: can't read data from url {_url}. {e}")
        sys.exit(3)
    print(f'Дата обновления данных на портале: {Color.DARKCYAN}{datetime.datetime.fromisoformat(data_json[0]["DATE_LAST_UPDATE"]).strftime("%d.%m.%Y %H:%M:%S")}{Color.END}')
    return data_json[0]['DATE_LAST_UPDATE']


def write_dataframe_to_file(_data_frame, _file_name, _sheet):
    """
    Записывает в Excel файл таблицы с данными

    :param _data_frame:
    :param _file_name:
    :param _sheet:
    """
    if Path(_file_name).is_file():
        with pd.ExcelWriter(_file_name, mode='a', if_sheet_exists="replace", datetime_format="DD.MM.YYYY",
                            engine='openpyxl') as writer:
            print(
                f'Append "{_sheet}" sheet to exist file: "{_file_name}"')
            _data_frame.to_excel(writer, sheet_name=_sheet, index=False)
    else:
        with pd.ExcelWriter(_file_name, mode='w', datetime_format="DD.MM.YYYY", engine='openpyxl') as writer:
            print(
                f'Write "{_sheet}" sheet to new file: "{_file_name}"')
            _data_frame.to_excel(writer, sheet_name=_sheet, index=False)


def convert_date(_data_frame, _columns):
    """
    Конвертирует поля с датами в формат datetime64.
    Возвращает конвертированный DataFrame

    :param _data_frame:
    :param _columns:
    :return DataFrame:
    """
    _columns_names = _data_frame.columns
    for _column_name in _columns_names:
        for _column in _columns:
            if _column.lower() in _column_name.lower():
                _data_frame[_column_name] = pd.to_datetime(_data_frame[_column_name], format='mixed', dayfirst=True, errors='ignore')  # , format="%d.%m.%Y"
            else:
                pass
    return _data_frame
    # for column in _columns:
    #     _data_frame[column] = pd.to_datetime(_data_frame[column], format='mixed', dayfirst=True, errors='ignore')  # , format="%d.%m.%Y"
    # return _data_frame


def convert_int(_data_frame, _columns):
    """
    Конвертирует поля с целыми в числовой формат.
    Возвращает конвертированный DataFrame

    :param _data_frame:
    :param _columns:
    :return DataFrame:
    """
    for column in _columns:
        _data_frame[column] = pd.to_numeric(_data_frame[column], errors='ignore')
    return _data_frame


def last_day_of_month(_date: datetime) -> datetime:
    if _date.month == 12:
        curr_year = _date.year + 1
        curr_month, curr_day = 1, 1
    else:
        curr_year = _date.year
        curr_month = _date.month + 1
        curr_day = 1
    return datetime.datetime(year=curr_year, month=curr_month, day=curr_day) - datetime.timedelta(microseconds=1)
    # return _date.replace(month=_date.month + 1, day=1) - datetime.timedelta(seconds=1)


def sum_sort_events(_data_frame, _column, _condition):
    _sum_sort = 0
    for _sum_data in _data_frame[_column]:
        if _sum_data in _condition:
            _sum_sort += 1
    return _sum_sort


def sum_done_events(_data_frame, _ks_date, _commissioning_date, _ks_status, _commissioning_status, _condition, _month, _last_days_of_month):
    _sum_sort = 0
    _sort_frame = _data_frame[[_ks_date, _commissioning_date, _ks_status, _commissioning_status]]
    for _row in _sort_frame.values:
        if pd.Timestamp(_row[0]) <= _last_days_of_month[_month] and pd.Timestamp(_row[1]) <= _last_days_of_month[_month] and _row[2] in _condition and _row[3] in _condition:
            _sum_sort += 1
    return _sum_sort


def sum_sort_month_events(_data_frame, _column, _month, _last_days_of_month):
    _sum_sort = 0
    for _sum_data in _data_frame[_column]:
        if pd.Timestamp(_sum_data) <= _last_days_of_month[_month]:
            _sum_sort += 1
    return _sum_sort


def adjust_columns_width(_dataframe):
    # Форматирование ширины полей отчётной таблицы
    for _col in _dataframe.columns:
        _max_length = 0
        _column = get_column_letter(_col[0].column)  # Get the column name
        for _cell in _col:
            if _cell.coordinate in _dataframe.merged_cells:  # not check merge_cells
                continue
            try:  # Necessary to avoid error on empty cells
                if len(str(_cell.value)) > _max_length:
                    _max_length = len(str(_cell.value))
            except Exception as e:
                logger.debug(f"Empty cell. Error text: {e}")
                pass
        _adjusted_width = _max_length + 3
        _dataframe.column_dimensions[_column].width = _adjusted_width
    return _dataframe


def megafon_send_email(data_frame: DataFrame, tag: str, template_directory: str, template_name: str, to_address: List[str], cc_address: List[str], attachment_file: bytes,
                       email_address: str, email_password: str, data_date: str):
    """
    @param data_frame:  Таблица DataFrame с данными
    @param tag:  Заголовок для формирования темы письма
    @param template_directory:  Директория с шаблонами писем
    @param template_name:  Наименование шаблона для письма
    @param to_address:  Список адресатов для письма
    @param cc_address:  Список адресатов для копии письма
    @param attachment_file: битовый массив c файлом Excel
    @param email_address: e-mail адрес от имени которого высылается рассылка
    @param email_password: Пароль для почтового сервера
    @param data_date: Дата обновления данных с портала
    """
    report_email = EmailSender(host='mail.megafon.ru', port=25, username=email_address, password=email_password, use_starttls=True)
    report_email.set_template_paths(html=Path(template_directory, 'html'))
    report_email.sender = email_address
    report_email.receivers = to_address
    if cc_address:
        report_email.cc = cc_address
    # Send the report
    print(f'Send email {Color.GREEN}"{tag}"{Color.END} to: {Color.GREEN}{to_address}{Color.END} and copy: {Color.GREEN}{cc_address}{Color.END}')
    report_email.send(
        subject=f'[A.M.S.] {tag}',
        html_template=template_name,
        body_params={
            'title': tag,
            'prog': PROGRAM_NAME,
            'ver': PROGRAM_VERSION,
            'data_date': datetime.datetime.fromisoformat(data_date).strftime("%d.%m.%Y %H:%M:%S"),
        },
        body_tables={"table": data_frame},
        attachments={
            f'{datetime.date.today().strftime("%Y%m%d")} {tag}.xlsx': attachment_file,
        }
    )


def call_send_email(dfs: DataFrame, email_list: list, no_debug: bool, email_address: str, email_password: str, last_update: str = None) -> None:
    my_email = email_address

    config = configparser.ConfigParser()
    try:
        with open(config_file, mode='r') as fr:
            config.read_file(fr)
    except FileNotFoundError:
        raise FileNotFoundError(f'Файл конфигурации {config_file} не найден.')

    config_dict = {}
    for key, value in config['MAILING_LISTS'].items():
        config_dict[key] = email_split(value)

    for key in ['cc_focl_no_tu', 'cc_focl_no_tu_to_po', 'cc_focl_tu_not_received_by_po']:
        config_dict[key].append(my_email)

    mailing_lists = {}
    for key in config_dict.keys():
        mailing_lists[key] = config_dict[key] if no_debug else my_email
    mailing_lists['me'] = my_email
    logger.debug(f'{mailing_lists = }')

    template_dir = 'templates'
    tag, tab_name, receivers, template = email_list
    mail_dfs = dfs.assign(пп=range(1, len(dfs) + 1)).set_index('пп').fillna('')
    to = mailing_lists[receivers[0]]
    cc = mailing_lists[receivers[1]]
    # Формируем временный файл с форматированной Excel таблицей для рассылки
    with BytesIO() as fp:
        logger.info(f'Создаем рабочую книгу для временного файла')
        mail_wb = FormattedWorkbook(properties_creator=email_address)
        mail_ws_first = mail_wb.active
        mail_wb.excel_format_table(dfs, tag, tab_name)
        logger.info(f'Удаляем лист {mail_ws_first}')
        mail_wb.remove(mail_ws_first)
        logger.info(f'Сохраняем временную книгу в {fp.__class__.__name__}')
        mail_wb.save(fp)
        temp_excel_file = fp.getvalue()

    megafon_send_email(mail_dfs, tag, template_dir, template, to, cc, temp_excel_file, email_address, email_password, last_update)


if __name__ == "__main__":
    pass
