#  Copyright (c) 2022. Tikhon Ostapenko
import argparse
import base64
import locale
import os
import threading

import openpyxl.styles.borders as borders_style
from dotenv import load_dotenv
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, Side, PatternFill, Alignment, Border

from vols_functions import *

# program and version
PROGRAM_NAME: str = "gdc_vols"
PROGRAM_VERSION: str = "0.6.28"


def main():
    # Константы
    BP = 'БП'
    BP_BUILD: str = 'Строительство ВОЛС'
    BP_RECON: str = 'Реконструкция ВОЛС'
    DELTA_CHAR = f'{chr(0x0394)}'

    load_dotenv()

    # Чтение переменных окружения
    EMAIL_ADDRESS = os.getenv('EMAIL_ADDRESS')
    EMAIL_PASSWORD = os.getenv('EMAIL_PASSWORD')
    if EMAIL_PASSWORD is not None:
        try:
            EMAIL_PASSWORD = base64.b64decode(base64.b85decode(EMAIL_PASSWORD.encode('UTF-8'))).decode('UTF-8')
        except ValueError:
            logger.error(f'Invalid email password')
            EMAIL_PASSWORD = None

    # Наименования колонок для преобразования даты
    columns_date = ['Планируемая дата окончания', 'Дата ввода', 'Прогнозная дата окончания', '_дата']
    # Наименования колонок для преобразования числа
    columns_digit = ['ID']
    # Наименование колонки для сортировки по возрастанию
    columns_for_sort = ['Регион/Зона мероприятия', 'Планируемая дата окончания']
    work_branch = "Кавказский филиал"
    today_date = datetime.date.today().strftime("%Y%m%d")  # YYYYMMDD format today date
    last_days_of_month = {}
    ext_build_df = None
    rec_df_ = None

    # Set local localization
    locale.setlocale(locale.LC_ALL, '')

    # Excel styles
    fn_bold = Font(bold=True)
    fn_red_bold = Font(color="FF0000", bold=True)
    fn_red = Font(color="B22222")
    fn_green = Font(color="006400")
    fn_green_bold = Font(color="006400", bold=True)
    fn_mag = Font(color="6633FF")
    # bd = Side(style='thick', color="000000")
    fill_red = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
    fill_yellow = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
    fill_green = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
    align_center = Alignment(horizontal="center")
    border_medium = Border(left=Side(style=borders_style.BORDER_MEDIUM), right=Side(style=borders_style.BORDER_MEDIUM),
                           top=Side(style=borders_style.BORDER_MEDIUM), bottom=Side(style=borders_style.BORDER_MEDIUM))
    border_thin = Border(left=Side(style=borders_style.BORDER_THIN), right=Side(style=borders_style.BORDER_THIN),
                         top=Side(style=borders_style.BORDER_THIN), bottom=Side(style=borders_style.BORDER_THIN))

    # Parse command line arguments
    parser = argparse.ArgumentParser(description=f'{PROGRAM_NAME} v.{PROGRAM_VERSION}')
    parser.add_argument("-v", "--verbose", type=int, help="Уровень отладки: 0 - CRITICAL, 1 - ERROR, 2 - INFO, 3 - DEBUG")
    parser.add_argument("-s", "--source-type", help="Тип источника данных (JSON или EXCEL)", default="JSON")
    parser.add_argument("-y", "--year", type=int, help="year for processing")
    parser.add_argument("-m", "--month", type=int, help="month for processing")
    parser.add_argument("-r", "--report-file", help="report file name, must have .xlsx extension")
    parser.add_argument("-b", "--report-branch", help="Branch name", default=work_branch)
    parser.add_argument("-l", "--send-email", action='store_true', help="Режим рассылки email")
    parser.add_argument("--no-debug", action='store_true', help="Запустить рассылку писем в \"боевом\" режиме")
    parser.add_argument("--new-algorithm", action='store_true', help="Использовать алгоритм подсчета по принятию в эксплуатацию, вместо факта КС-2 и вводу в эксплуатацию")
    parser.add_argument("--active-year", action='store_true', help="Формировать список активных мероприятий до конца года")
    parser.add_argument("--soc-report", action='store_true', help="Добавить в отчет страницы Соц. соревнования")
    parser.add_argument("--ignore-cert", action='store_true', help="Игнорировать проверку SSL сертификатов при получении данных")
    parser.add_argument("--no-update-date", action='store_true', help="Не запрашивать дату обновления с портала")
    args = parser.parse_args()

    # Добавление суффикса к имени сохраняемого файла при задании режимов работы
    file_suffix = ''
    if any([args.new_algorithm, args.soc_report, args.active_year]):
        file_suffix = f'{" (new-algorithm)" if args.new_algorithm else ""}{" (soc-report)" if args.soc_report else ""}{" (active-year])" if args.active_year else ""}'

    # Уровень отладочных сообщений
    if args.verbose is None or args.verbose == 1:
        logger_level = 'ERROR'
    elif args.verbose == 0:
        logger_level = 'CRITICAL'
    elif args.verbose == 2:
        logger_level = 'INFO'
    else:
        logger_level = 'DEBUG'
    logger.remove()
    logger.add(sys.stdout, level=logger_level)

    # Год анализа.
    if args.year is None:
        process_year = datetime.date.today().year
    else:
        process_year = args.year

        # Месяц для анализа.
    if args.month is None:
        process_month = datetime.date.today().month
    else:
        process_month = args.month

    if args.report_branch is not None:
        work_branch = args.report_branch

    if args.report_file is None:
        vols_dir = Path('//megafon.ru/KVK/KRN/Files/TelegrafFiles/ОПРС/!Проекты РЦРП/Блок №3/ВОЛС', str(process_year))
        vols_file = f'{today_date} Отчет по строительству и реконструкции ВОЛС {"".join(symbol[0].upper() for symbol in work_branch.split())} {datetime.date(process_year, process_month, 1).strftime("%m.%Y")}{file_suffix}.xlsx'
        file_name = Path(vols_dir, vols_file)
    else:
        if Path(args.report_file).parent.exists():
            file_name = Path(args.report_file)
            if Path(args.report_file).suffix != '.xlsx':
                file_name = file_name.with_suffix('.xlsx')
        else:
            logger.error(f'Директория для файла отчета {Path(args.report_file).parent} не существует')
            sys.exit(100)

    if args.ignore_cert:
        check_cert = False
    else:
        check_cert = True

    api_urls = {
        # f'Расш. стр. гор.ВОЛС {process_year}': f'https://vlg-adi-web01.megafon.ru/legacy-dash/dashboard/plan/vw_{process_year}_FOCL_Common_Build_City_211_dev',
        f'Расш. стр. гор.ВОЛС {process_year}': f'https://vlg-adi-web01.megafon.ru/legacy-dash/dashboard/plan/vw_{process_year}_FOCL_Common_Build_City',
        f'Cтр. гор.ВОЛС (РАП) {process_year}': f'https://vlg-adi-web01.megafon.ru/legacy-dash/dashboard/plan/vw_{process_year + 1}_FOCL_Common_Build_City',
        f'Реконструкция гор.ВОЛС {process_year}': f'https://vlg-adi-web01.megafon.ru/legacy-dash/dashboard/plan/vw_{process_year}_FOCL_Common_Rebuild_City',
        f'Строительство зон.ВОЛС {process_year}': f'https://vlg-adi-web01.megafon.ru/legacy-dash/dashboard/plan/vw_{process_year}_FOCL_Common_Build_Zone',
        f'Реконструкция зон.ВОЛС {process_year}': f'https://vlg-adi-web01.megafon.ru/legacy-dash/dashboard/plan/vw_{process_year}_FOCL_Common_Rebuild_Zone',
    }

    excel_urls = {
        f'Расш. стр. гор.ВОЛС {process_year}': f'https://gdc-rts.megafon.ru/api/legacy/download?table=vw_{process_year}_FOCL_Common_Build_City&database=dashboard',
        f'Cтр. гор.ВОЛС (РАП) {process_year}': f'https://gdc-rts.megafon.ru/api/legacy/download?table=vw_{process_year + 1}_FOCL_Common_Build_City&database=dashboard',
        f'Реконструкция гор.ВОЛС {process_year}': f'https://gdc-rts.megafon.ru/api/legacy/download?table=vw_{process_year}_FOCL_Common_Rebuild_City&database=dashboard',
        f'Строительство зон.ВОЛС {process_year}': f'https://gdc-rts.megafon.ru/api/legacy/download?table=vw_{process_year}_FOCL_Common_Build_Zone&database=dashboard',
        f'Реконструкция зон.ВОЛС {process_year}': f'https://gdc-rts.megafon.ru/api/legacy/download?table=vw_{process_year}_FOCL_Common_Rebuild_Zone&database=dashboard',
    }

    last_update_url = f"https://vlg-adi-web01.megafon.ru/legacy-dash/dashboard/upd/fn_{process_year}_FOCL_Plan_Build_City()"

    data_sheets = {
        'city_main_build': f'Осн. стр. гор.ВОЛС {process_year}',
        'city_ext_build': f'Доп. стр. гор.ВОЛС {process_year}',
        'city_reconstruction': f'Реконструкция гор.ВОЛС {process_year}',
        'zone_build': f'Строительство зон.ВОЛС {process_year}',
        'zone_reconstruction': f'Реконструкция зон.ВОЛС {process_year}'
    }

    report_sheets = {
        'report': "Отчетная таблица",
        'current_month': f'Активные мероприятия {datetime.date(process_year, process_month, 1).strftime("%m.%Y") if not args.active_year else process_year}',
        'tz': 'Нет ТЗ',
        'sending_po': "Нет передачи ТЗ в ПО",
        'received_po': 'ПО не приняли ТЗ в ЕСУП',
        'soc_build': 'Соц.соревнование. Стр.',
        'soc_rec': 'Соц. соревнование. Рек.',
    }

    reports_data = {
        'tz': [f'ВОЛС. {report_sheets["tz"]}', "FOCL_no_TU", ['focl_no_tu', 'cc_focl_no_tu'], 'focl_no_tu.html'],
        'sending_po': [f'ВОЛС. {report_sheets["sending_po"]}', "FOCL_no_TU_to_PO", ['focl_no_tu_to_po', 'cc_focl_no_tu_to_po'], 'focl_no_tu_to_po.html'],
        'received_po': [f'ВОЛС. {report_sheets["received_po"]}', "FOCL_TU_not_received_by_PO", ['focl_tu_not_received_by_po', 'cc_focl_tu_not_received_by_po'],
                        'focl_tu_not_received_by_po.html'],
    }

    excel_tables_names = {
        data_sheets['city_main_build']: "Urban_VOLS_Main_Build",
        data_sheets['city_ext_build']: "Urban_VOLS_Ext_Build",
        data_sheets['city_reconstruction']: "Urban_VOLS_Reconstruction",
        data_sheets['zone_build']: "Zone_VOLS_Build",
        data_sheets['zone_reconstruction']: "Zone_VOLS_Reconstruction",
        report_sheets['current_month']: "current_month",
        report_sheets['tz']: "tz_not_done",
        report_sheets['sending_po']: "sending_po_not_done",
        report_sheets['received_po']: "received_po_not_done",
        report_sheets['soc_build']: "soc_build",
        report_sheets['soc_rec']: "soc_rec",
    }

    # excel_cell_names = fill_cell_names()

    process_columns = {
        'plan_date': 'Планируемая дата окончания',
        'tz_date': 'Разработка ТЗ_дата',
        'tz_date2': 'Разработка ТЗ ВОЛС_Дата',
        'send_tz_date': 'Передача ТЗ подрядчику_дата',
        'send_tz_date2': 'Передача ТЗ на ВОЛС подрядчику_Дата',
        'received_tz_date': 'ТЗ принято подрядчиком_дата',
        'received_tz_date2': 'ТЗ принято подрядчиком_дата',
        'pir_smr_date': 'Заказ ПИР,СМР_дата',
        'pir_smr_date2': 'Подписание договора (дс/заказа) на ПИР/ПИР+СМР_Дата',
        'line_scheme_date': 'Линейная схема_дата',
        'line_scheme_date2': 'Линейная схема_Дата',
        'tu_date': 'Получение ТУ_дата',
        'tu_date2': 'Получение ТУ_Дата',
        'build_date': 'Строительство трассы_дата',
        'build_date2': 'Строительство трассы_Дата',
        'ks2_date': 'КС-2 (ПИР, СМР)_дата',
        'ks2_date2': 'КС-2,3_Дата',
        'commissioning_date': 'Приемка в эксплуатацию_дата',
        'commissioning_date2': 'Приемка ВОЛС в эксплуатацию_Дата',
        'complete_date': 'Дата ввода в эксплуатацию',
        'complete_date2': 'Дата ввода ВОЛС в эксплуатацию',
        'traffic_date': 'Запуск трафика_дата',
        'traffic_date2': 'Запуск трафика_Дата',
        'tz_status': 'Разработка ТЗ_статус',
        'tz_status2': 'Разработка ТЗ ВОЛС_Статус',
        'send_tz_status': 'Передача ТЗ подрядчику_статус',
        'send_tz_status2': 'Передача ТЗ на ВОЛС подрядчику_Статус',
        'received_tz_status': 'ТЗ принято подрядчиком_статус',
        'received_tz_status2': 'ТЗ принято подрядчиком_статус',
        'pir_smr_status': 'Заказ ПИР,СМР_статус',
        'pir_smr_status2': 'Подписание договора (дс/заказа) на ПИР/ПИР+СМР_Статус',
        'line_scheme_status': 'Линейная схема_статус',
        'line_scheme_status2': 'Линейная схема_Статус',
        'line_scheme_status3': 'Линейная схема (АВТ)_статус',
        'tu_status': 'Получение ТУ_статус',
        'tu_status2': 'Получение ТУ_Статус',
        'build_status': 'Строительство трассы_статус',
        'build_status2': 'Строительство трассы_Статус',
        'ks2_status': 'КС-2 (ПИР, СМР)_статус',
        'ks2_status2': 'КС-2,3_Статус',
        'commissioning_status': 'Приемка в эксплуатацию_статус',
        'commissioning_status2': 'Приемка ВОЛС в эксплуатацию_Статус',
        'traffic_status': 'Запуск трафика_статус',
        'traffic_status2': 'Запуск трафика_Статус',
        'id': 'ID',
        'branch': 'Филиал',
        'region': 'Регион/Зона мероприятия',
        'name': 'Название',
        'program': 'Программы',
        'prognoz_date': 'Прогнозная дата окончания',
    }

    rename_columns = {
        process_columns['complete_date2']: process_columns['complete_date'],
    }
    soc_rename_columns = {
        process_columns['plan_date']: 'План',
        process_columns['complete_date']: 'Факт',
    }

    print(f'{Color.DARKCYAN}{datetime.datetime.now().strftime("%d.%m.%Y %H:%M:%S")}:{Color.END} {PROGRAM_NAME}: {PROGRAM_VERSION}')

    wb = FormattedWorkbook(logging_level=logger_level, properties_creator=EMAIL_ADDRESS)
    ws_first = wb.active

    # Получение исходных данных и запись форматированных данных

    # Временно выключаем проверку сертификатов
    ssl._create_default_https_context = ssl._create_unverified_context
    # Временно выключаем проверку сертификатов

    # Получаем дату обновления данных на портале
    date_last_update = None
    # date_last_update = datetime.datetime.now().isoformat()
    if not args.no_update_date:
        date_last_update = get_update_date(last_update_url, check_ssl=check_cert)
        if date_last_update is not None:
            data_update_age = (datetime.datetime.now() - datetime.datetime.fromisoformat(date_last_update))
            if data_update_age > datetime.timedelta(hours=48):
                if input(
                        f'{Color.RED}Данные на портале обновлялись {data_update_age.days * 24 + data_update_age.seconds / 3600:.2f} час. назад! Хотите продолжить обработку данных (y/N)?{Color.END}').lower() != 'y':
                    sys.exit(12)

    # Получаем данные с портала
    if args.source_type.lower() == "excel":
        urls = excel_urls  # Скачиваем EXCEL файлы
        input_data_type = "EXCEL"
    else:
        urls = api_urls  # Скачиваем по API JSON
        input_data_type = "JSON"

    for sheet, url in urls.items():
        data_frame = read_from_dashboard(url, data_type=input_data_type, check_ssl=check_cert)  # Читаем данные из сети. Для API запросов data_type должен быть "JSON", для скачиваемых файлов "EXCEL"
        if process_columns['branch'] in data_frame.columns:
            data_frame = data_frame[data_frame[process_columns['branch']] == work_branch]  # Оставляем только отчётный филиал
        else:
            print(f'{Color.RED}Не корректный формат входящих данных. Проверьте корректность данных для {sheet}!{Color.END}')
            sys.exit(2)

        # Переименовываем столбцы в таблицах
        if process_columns['line_scheme_status3'] in data_frame.columns:
            data_frame.rename(columns={process_columns['line_scheme_status3']: process_columns['line_scheme_status']}, inplace=True)

        data_frame = data_frame.reset_index(drop=True)
        data_frame = convert_date(data_frame, columns_date)  # Переводим дату в формат datetime
        data_frame = convert_int(data_frame, columns_digit)  # Переводим ESUP_ID в числовой формат
        data_frame = data_frame.sort_values(by=columns_for_sort)  # Сортируем по заданному столбцу
        if sheet == f'Расш. стр. гор.ВОЛС {process_year}':
            extended_build_df = data_frame.copy(deep=True)  # keep extended data for analyses
            # Формируем таблицу основного строительства
            # main_build_df = data_frame[data_frame['KPI ПТР текущего года, км'].notna() & (data_frame['KPI ПТР текущего года, км'] > 0)]
            main_build_df = data_frame
            if not main_build_df.empty:
                print(f'Создаем лист: {Color.GREEN}"{data_sheets["city_main_build"]}"{Color.END}')
                wb.excel_format_table(
                    main_build_df,
                    data_sheets['city_main_build'],
                    excel_tables_names[data_sheets['city_main_build']],
                )
            # Формируем таблицу дополнительного строительства
            # ext_build_df = data_frame[~data_frame['KPI ПТР текущего года, км'].notna() | (data_frame['KPI ПТР текущего года, км'] == 0)]
        elif sheet == f'Cтр. гор.ВОЛС (РАП) {process_year}':
            ext_build_df = data_frame
            if not ext_build_df.empty:
                print(f'Создаем лист: {Color.GREEN}"{data_sheets["city_ext_build"]}"{Color.END}')
                wb.excel_format_table(
                    ext_build_df,
                    data_sheets['city_ext_build'],
                    excel_tables_names[data_sheets['city_ext_build']],
                )
        else:
            if sheet == f'Реконструкция гор.ВОЛС {process_year}':
                rec_df_ = data_frame
            if not data_frame.empty:
                print(f'Создаем лист: {Color.GREEN}"{sheet}"{Color.END}')
                wb.excel_format_table(
                    data_frame,
                    sheet,
                    excel_tables_names[sheet],
                )

    # Создание отчёта
    print(f'Создаем лист отчета: {Color.GREEN}"{report_sheets["report"]}"{Color.END}')
    for i in range(1, 13):
        last_days_of_month[i] = pd.Timestamp(last_day_of_month(datetime.datetime(process_year, i, 1)))
    try:
        ws = wb[report_sheets['report']]
    except KeyError:
        logger.info(f"Лист \"{report_sheets['report']}\" не существует. Создаем... ")
        ws = wb.create_sheet(title=report_sheets['report'], index=0)

    # Формирование статических полей отчёта
    if date_last_update is not None:
        ws['K1'] = "Дата обновления данных"
        ws['L1'] = datetime.datetime.fromisoformat(date_last_update).strftime("%d.%m.%Y %H:%M:%S")
        ws['L1'].font = fn_green_bold
        ws['L1'].alignment = align_center
    ws['A1'] = "Основное строительство ВОЛС"
    ws['A1'].font = fn_red_bold
    ws['A1'].border = border_thin
    ws['A2'] = 'Всего мероприятий'
    ws['A2'].border = border_medium
    ws['A4'] = 'Исполнение KPI ВОЛС КФ (накопительный итог)'
    ws['A4'].font = fn_red_bold
    ws['A4'].border = border_thin
    ws['A6'] = 'Учтенных ВОЛС в KPI'
    ws['A6'].border = border_medium
    ws['A8'] = 'Исполнение мероприятий в ЕСУП'
    ws['A8'].font = fn_red_bold
    ws['A8'].border = border_thin
    ws['A9'] = 'Наименование мероприятия'
    ws['A9'].font = fn_bold
    ws['A9'].border = border_medium
    ws['A10'] = 'Выпущены ТЗ'
    ws['A10'].border = border_medium
    ws['A11'] = 'Переданы ТЗ в ПО'
    ws['A11'].border = border_medium
    ws['A12'] = 'Приняты ТЗ ПО'
    ws['A12'].border = border_medium
    ws['A13'] = 'Подписание договора на ПИР/ПИР+СМР'
    ws['A13'].border = border_medium
    ws['A14'] = 'Линейная схема'
    ws['A14'].border = border_medium
    ws['A15'] = 'Получено ТУ'
    ws['A15'].border = border_medium
    ws['A16'] = 'Строительство трассы'
    ws['A16'].border = border_medium
    ws['A17'] = 'Подготовка актов КС-2,3'
    ws['A17'].border = border_medium
    ws['A18'] = 'Приёмка ВОЛС в эксплуатацию'
    ws['A18'].border = border_medium
    ws['B9'] = 'Выполнено'
    ws['B9'].font = fn_bold
    ws['B9'].alignment = align_center
    ws['B9'].border = border_medium
    ws['C9'] = DELTA_CHAR
    ws['C9'].font = fn_bold
    ws['C9'].alignment = align_center
    ws['C9'].border = border_medium

    if ext_build_df is not None:
        ws['A21'] = "Дополнительное строительство ВОЛС"
        ws['A21'].font = fn_red_bold
        ws['A21'].border = border_thin
        ws['A22'] = 'Всего мероприятий'
        ws['A22'].border = border_medium
        ws['A24'] = 'Исполнение KPI ВОЛС КФ (накопительный итог)'
        ws['A24'].font = fn_red_bold
        ws['A24'].border = border_thin
        ws['A26'] = 'Учтенных ВОЛС в KPI'
        ws['A26'].border = border_medium
        ws['A28'] = 'Исполнение мероприятий в ЕСУП'
        ws['A28'].font = fn_red_bold
        ws['A28'].border = border_thin
        ws['A29'] = 'Наименование мероприятия'
        ws['A29'].font = fn_bold
        ws['A29'].border = border_medium
        ws['A30'] = 'Выпущены ТЗ'
        ws['A30'].border = border_medium
        ws['A31'] = 'Переданы ТЗ в ПО'
        ws['A31'].border = border_medium
        ws['A32'] = 'Приняты ТЗ ПО'
        ws['A32'].border = border_medium
        ws['A33'] = 'Подписание договора на ПИР/ПИР+СМР'
        ws['A33'].border = border_medium
        ws['A34'] = 'Линейная схема'
        ws['A34'].border = border_medium
        ws['A35'] = 'Получено ТУ'
        ws['A35'].border = border_medium
        ws['A36'] = 'Строительство трассы'
        ws['A36'].border = border_medium
        ws['A37'] = 'Подготовка актов КС-2,3'
        ws['A37'].border = border_medium
        ws['A38'] = 'Приёмка ВОЛС в эксплуатацию'
        ws['A38'].border = border_medium
        ws['B29'] = 'Выполнено'
        ws['B29'].font = fn_bold
        ws['B29'].alignment = align_center
        ws['B29'].border = border_medium
        ws['C29'] = DELTA_CHAR
        ws['C29'].font = fn_bold
        ws['C29'].alignment = align_center
        ws['C29'].border = border_medium

    if rec_df_ is not None:
        ws['F1'] = "Реконструкция ВОЛС"
        ws['F1'].font = fn_red_bold
        ws['F1'].border = border_thin
        ws['F2'] = 'Всего мероприятий'
        ws['F2'].border = border_medium
        ws['F2'].border = border_medium
        ws['F4'] = 'Исполнение KPI ВОЛС КФ (накопительный итог)'
        ws['F4'].font = fn_red_bold
        ws['F4'].border = border_thin
        ws['F6'] = 'Учтенных ВОЛС в KPI'
        ws['F6'].border = border_medium
        ws['F8'] = 'Исполнение мероприятий в ЕСУП'
        ws['F8'].font = fn_red_bold
        ws['F8'].border = border_thin
        ws['F9'] = 'Наименование мероприятия'
        ws['F9'].font = fn_bold
        ws['F9'].border = border_medium
        ws['F10'] = 'Выпущены ТЗ'
        ws['F10'].border = border_medium
        ws['F11'] = 'Переданы ТЗ в ПО'
        ws['F11'].border = border_medium
        ws['F12'] = 'Приняты ТЗ ПО'
        ws['F12'].border = border_medium
        ws['F13'] = 'Подписание договора на ПИР/ПИР+СМР'
        ws['F13'].border = border_medium
        ws['F14'] = 'Линейная схема'
        ws['F14'].border = border_medium
        ws['F15'] = 'Получено ТУ'
        ws['F15'].border = border_medium
        ws['F16'] = 'Строительство трассы'
        ws['F16'].border = border_medium
        ws['F17'] = 'Подготовка актов КС-2,3'
        ws['F17'].border = border_medium
        ws['F18'] = 'Приёмка ВОЛС в эксплуатацию'
        ws['F18'].border = border_medium
        ws['G9'] = 'Выполнено'
        ws['G9'].font = fn_bold
        ws['G9'].alignment = align_center
        ws['G9'].border = border_medium
        ws['H9'] = DELTA_CHAR
        ws['H9'].font = fn_bold
        ws['H9'].alignment = align_center
        ws['H9'].border = border_medium

    # Поля для целевых мероприятий по Base Case
    ws['A41'] = "Целевые мероприятия Base Case ВОЛС"
    ws['A41'].font = fn_red_bold
    ws['A41'].border = border_thin
    ws['A42'] = 'Всего мероприятий'
    ws['A42'].border = border_medium
    ws['A44'] = 'Исполнение KPI Base Case (накопительный итог)'
    ws['A44'].font = fn_red_bold
    ws['A44'].border = border_thin
    ws['A46'] = 'Учтенных ВОЛС в KPI'
    ws['A46'].border = border_medium
    ws['A48'] = 'Исполнение мероприятий в ЕСУП'
    ws['A48'].font = fn_red_bold
    ws['A48'].border = border_thin
    ws['A49'] = 'Наименование мероприятия'
    ws['A49'].font = fn_bold
    ws['A49'].border = border_medium
    ws['A50'] = 'Выпущены ТЗ'
    ws['A50'].border = border_medium
    ws['A51'] = 'Переданы ТЗ в ПО'
    ws['A51'].border = border_medium
    ws['A52'] = 'Приняты ТЗ ПО'
    ws['A52'].border = border_medium
    ws['A53'] = 'Подписание договора на ПИР/ПИР+СМР'
    ws['A53'].border = border_medium
    ws['A54'] = 'Линейная схема'
    ws['A54'].border = border_medium
    ws['A55'] = 'Получено ТУ'
    ws['A55'].border = border_medium
    ws['A56'] = 'Строительство трассы'
    ws['A56'].border = border_medium
    ws['A57'] = 'Подготовка актов КС-2,3'
    ws['A57'].border = border_medium
    ws['A58'] = 'Приёмка ВОЛС в эксплуатацию'
    ws['A58'].border = border_medium
    ws['B49'] = 'Выполнено'
    ws['B49'].font = fn_bold
    ws['B49'].alignment = align_center
    ws['B49'].border = border_medium
    ws['C49'] = DELTA_CHAR
    ws['C49'].font = fn_bold
    ws['C49'].alignment = align_center
    ws['C49'].border = border_medium

    # Формирование динамических полей отчёта
    ws['B5'] = f'План, {datetime.date(process_year, process_month, 1).strftime("%b %Y")}'
    ws['B5'].font = fn_bold
    ws['B5'].alignment = align_center
    ws['B5'].border = border_medium
    ws['C5'] = f'Факт, {datetime.date(process_year, process_month, 1).strftime("%b %Y")}'
    ws['C5'].font = fn_bold
    ws['C5'].alignment = align_center
    ws['C5'].border = border_medium
    ws['D5'] = f'{chr(0x0394)}, {datetime.date(process_year, process_month, 1).strftime("%b %Y")}'
    ws['D5'].font = fn_bold
    ws['D5'].alignment = align_center
    ws['D5'].border = border_medium

    if ext_build_df is not None:
        ws['B25'] = f'План, {datetime.date(process_year, process_month, 1).strftime("%b %Y")}'
        ws['B25'].font = fn_bold
        ws['B25'].alignment = align_center
        ws['B25'].border = border_medium
        ws['C25'] = f'Факт, {datetime.date(process_year, process_month, 1).strftime("%b %Y")}'
        ws['C25'].font = fn_bold
        ws['C25'].alignment = align_center
        ws['C25'].border = border_medium
        ws['D25'] = f'{chr(0x0394)}, {datetime.date(process_year, process_month, 1).strftime("%b %Y")}'
        ws['D25'].font = fn_bold
        ws['D25'].alignment = align_center
        ws['D25'].border = border_medium

    if rec_df_ is not None:
        ws['G5'] = f'План, {datetime.date(process_year, process_month, 1).strftime("%b %Y")}'
        ws['G5'].font = fn_bold
        ws['G5'].alignment = align_center
        ws['G5'].border = border_medium
        ws['H5'] = f'Факт, {datetime.date(process_year, process_month, 1).strftime("%b %Y")}'
        ws['H5'].font = fn_bold
        ws['H5'].alignment = align_center
        ws['H5'].border = border_medium
        ws['I5'] = f'{chr(0x0394)}, {datetime.date(process_year, process_month, 1).strftime("%b %Y")}'
        ws['I5'].font = fn_bold
        ws['I5'].alignment = align_center
        ws['I5'].border = border_medium

    ws['B45'] = f'План, {datetime.date(process_year, process_month, 1).strftime("%b %Y")}'
    ws['B45'].font = fn_bold
    ws['B45'].alignment = align_center
    ws['B45'].border = border_medium
    ws['C45'] = f'Факт, {datetime.date(process_year, process_month, 1).strftime("%b %Y")}'
    ws['C45'].font = fn_bold
    ws['C45'].alignment = align_center
    ws['C45'].border = border_medium
    ws['D45'] = f'{chr(0x0394)}, {datetime.date(process_year, process_month, 1).strftime("%b %Y")}'
    ws['D45'].font = fn_bold
    ws['D45'].alignment = align_center
    ws['D45'].border = border_medium

    # Анализ строительства ВОЛС
    # TODO: Необходимо переделать генерацию отчетной страницы на процедуры или классы
    df = extended_build_df.copy(deep=True)

    # main_build_df = df[df['KPI ПТР текущего года, км'].notnull() & (df['KPI ПТР текущего года, км'] > 0)]
    # main_build_df = df
    # ext_build_df = df

    if ext_build_df is not None:
        df = pd.concat([main_build_df, ext_build_df], ignore_index=True).reset_index(drop=True)

    kpi_build_df = main_build_df[main_build_df[process_columns['program']].str.match(r'.*Base Case.*')]

    build_dashboard_data = df.copy(deep=True)
    tz_build_dataframe = df[df[process_columns['tz_status']] != 'Исполнена']
    sending_po_build_dataframe = df[df[process_columns['send_tz_status']] != 'Исполнена']
    received_po_build_dataframe = df[df[process_columns['received_tz_status']] != 'Исполнена']

    ws['B2'] = main_build_df[process_columns['plan_date']].count()
    ws['B2'].font = fn_bold
    ws['B2'].alignment = align_center
    ws['B2'].border = border_medium
    # ws['B6'] = sum_sort_month_events(main_build_df, process_columns['plan_date'], process_month, last_days_of_month)
    ws['B6'] = main_build_df[(main_build_df[process_columns['plan_date']] != '') & (
            main_build_df[process_columns['plan_date']] <= last_days_of_month[process_month])][process_columns['plan_date']].count()
    ws['B6'].alignment = align_center
    ws['B6'].border = border_medium

    if not args.new_algorithm:
        ws['C6'] = main_build_df[(main_build_df[process_columns['commissioning_date']] != '') & (
                main_build_df[process_columns['commissioning_date']] <= last_days_of_month[process_month]) & (main_build_df[process_columns['ks2_date']] != '') & (
                                         main_build_df[process_columns['ks2_date']] <= last_days_of_month[process_month])][process_columns['commissioning_date']].count()
    else:
        ws['C6'] = main_build_df[(main_build_df[process_columns['complete_date']] != '') & (
                main_build_df[process_columns['complete_date']] <= last_days_of_month[process_month])][process_columns['complete_date']].count()
    ws['C6'].alignment = align_center
    ws['C6'].border = border_medium
    ws['D6'] = ws['C6'].value - ws['B6'].value
    ws['D6'].alignment = align_center
    ws['D6'].border = border_medium
    ws.conditional_formatting.add('D6', CellIsRule(operator='lessThan', formula=['0'], stopIfTrue=True, font=fn_red, fill=fill_red))
    ws.conditional_formatting.add('D6', CellIsRule(operator='greaterThan', formula=['0'], stopIfTrue=True, font=fn_green, fill=fill_green))
    ws.conditional_formatting.add('D6', CellIsRule(operator='equal', formula=['0'], stopIfTrue=True, font=fn_mag, fill=fill_yellow))

    if ext_build_df is not None:
        ws['B22'] = ext_build_df[process_columns['plan_date']].count()
        ws['B22'].font = fn_bold
        ws['B22'].alignment = align_center
        ws['B22'].border = border_medium
        # ws['B26'] = sum_sort_month_events(ext_build_df, process_columns['plan_date'], process_month, last_days_of_month)
        ws['B26'] = ext_build_df[(ext_build_df[process_columns['plan_date']] != '') & (
                ext_build_df[process_columns['plan_date']] <= last_days_of_month[process_month])][process_columns['plan_date']].count()
        ws['B26'].alignment = align_center
        ws['B26'].border = border_medium
        if not args.new_algorithm:
            ws['C26'] = ext_build_df[
                (ext_build_df[process_columns['commissioning_date']] != '') & (ext_build_df[process_columns['commissioning_date']] <= last_days_of_month[process_month]) & (
                        ext_build_df[process_columns['ks2_date']] != '') & (ext_build_df[process_columns['ks2_date']] <= last_days_of_month[process_month])][
                process_columns['commissioning_date']].count()
        else:
            ws['C26'] = ext_build_df[(ext_build_df[process_columns['complete_date']] != '') & (
                    ext_build_df[process_columns['complete_date']] <= last_days_of_month[process_month])][process_columns['complete_date']].count()
        ws['C26'].alignment = align_center
        ws['C26'].border = border_medium
        ws['D26'] = ws['C26'].value - ws['B26'].value
        ws['D26'].alignment = align_center
        ws['D26'].border = border_medium
        ws.conditional_formatting.add('D26', CellIsRule(operator='lessThan', formula=['0'], stopIfTrue=True, font=fn_red, fill=fill_red))
        ws.conditional_formatting.add('D26', CellIsRule(operator='greaterThan', formula=['0'], stopIfTrue=True, font=fn_green, fill=fill_green))
        ws.conditional_formatting.add('D26', CellIsRule(operator='equal', formula=['0'], stopIfTrue=True, font=fn_mag, fill=fill_yellow))

    ws['B42'] = kpi_build_df[process_columns['plan_date']].count()
    ws['B42'].font = fn_bold
    ws['B42'].alignment = align_center
    ws['B42'].border = border_medium
    # ws['B46'] = sum_sort_month_events(kpi_build_df, process_columns['plan_date'], process_month, last_days_of_month)
    ws['B46'] = kpi_build_df[(kpi_build_df[process_columns['plan_date']] != '') & (
            kpi_build_df[process_columns['plan_date']] <= last_days_of_month[process_month])][process_columns['plan_date']].count()
    ws['B46'].alignment = align_center
    ws['B46'].border = border_medium
    if not args.new_algorithm:
        ws['C46'] = kpi_build_df[
            (kpi_build_df[process_columns['commissioning_date']] != '') & (kpi_build_df[process_columns['commissioning_date']] <= last_days_of_month[process_month]) & (
                    kpi_build_df[process_columns['ks2_date']] != '') & (kpi_build_df[process_columns['ks2_date']] <= last_days_of_month[process_month])][
            process_columns['commissioning_date']].count()
    else:
        ws['C46'] = kpi_build_df[(kpi_build_df[process_columns['complete_date']] != '') & (
                kpi_build_df[process_columns['complete_date']] <= last_days_of_month[process_month])][process_columns['complete_date']].count()
    ws['C46'].alignment = align_center
    ws['C46'].border = border_medium
    ws['D46'] = ws['C46'].value - ws['B46'].value
    ws['D46'].alignment = align_center
    ws['D46'].border = border_medium
    ws.conditional_formatting.add('D46', CellIsRule(operator='lessThan', formula=['0'], stopIfTrue=True, font=fn_red, fill=fill_red))
    ws.conditional_formatting.add('D46', CellIsRule(operator='greaterThan', formula=['0'], stopIfTrue=True, font=fn_green, fill=fill_green))
    ws.conditional_formatting.add('D46', CellIsRule(operator='equal', formula=['0'], stopIfTrue=True, font=fn_mag, fill=fill_yellow))

    for i, process in zip(range(10, 19), ['tz_status',
                                          'send_tz_status',
                                          'received_tz_status',
                                          'pir_smr_status',
                                          'line_scheme_status',
                                          'tu_status',
                                          'build_status',
                                          'ks2_status',
                                          'commissioning_status',
                                          ]):
        ws[f'B{i}'] = sum_sort_events(main_build_df, process_columns[process], ['Исполнена', 'Не требуется'])
        ws[f'B{i}'].alignment = align_center
        ws[f'B{i}'].border = border_medium
    for i in range(10, 19):
        ws[f'C{i}'] = ws[f'B{i}'].value - ws['B2'].value
        ws[f'C{i}'].alignment = align_center
        ws[f'C{i}'].border = border_medium
        ws.conditional_formatting.add(f'C{i}', CellIsRule(operator='greaterThanOrEqual', formula=['0'], stopIfTrue=True, font=fn_green, fill=fill_green))
        ws.conditional_formatting.add(f'C{i}', CellIsRule(operator='lessThan', formula=['0'], stopIfTrue=True, font=fn_red, fill=fill_red))

        if ext_build_df is not None:
            for i, process in zip(range(30, 39), ['tz_status',
                                                  'send_tz_status',
                                                  'received_tz_status',
                                                  'pir_smr_status',
                                                  'line_scheme_status',
                                                  'tu_status',
                                                  'build_status',
                                                  'ks2_status',
                                                  'commissioning_status',
                                                  ]):
                ws[f'B{i}'] = sum_sort_events(ext_build_df, process_columns[process], ['Исполнена', 'Не требуется'])
                ws[f'B{i}'].alignment = align_center
                ws[f'B{i}'].border = border_medium
            for i in range(30, 39):
                ws[f'C{i}'] = ws[f'B{i}'].value - ws['B22'].value
                ws[f'C{i}'].alignment = align_center
                ws[f'C{i}'].border = border_medium
                ws.conditional_formatting.add(f'C{i}', CellIsRule(operator='greaterThanOrEqual', formula=['0'], stopIfTrue=True, font=fn_green, fill=fill_green))
                ws.conditional_formatting.add(f'C{i}', CellIsRule(operator='lessThan', formula=['0'], stopIfTrue=True, font=fn_red, fill=fill_red))

    for i, process in zip(range(50, 59), ['tz_status',
                                          'send_tz_status',
                                          'received_tz_status',
                                          'pir_smr_status',
                                          'line_scheme_status',
                                          'tu_status',
                                          'build_status',
                                          'ks2_status',
                                          'commissioning_status',
                                          ]):
        ws[f'B{i}'] = sum_sort_events(kpi_build_df, process_columns[process], ['Исполнена', 'Не требуется'])
        ws[f'B{i}'].alignment = align_center
        ws[f'B{i}'].border = border_medium
    for i in range(50, 59):
        ws[f'C{i}'] = ws[f'B{i}'].value - ws['B42'].value
        ws[f'C{i}'].alignment = align_center
        ws[f'C{i}'].border = border_medium
        ws.conditional_formatting.add(f'C{i}', CellIsRule(operator='greaterThanOrEqual', formula=['0'], stopIfTrue=True, font=fn_green, fill=fill_green))
        ws.conditional_formatting.add(f'C{i}', CellIsRule(operator='lessThan', formula=['0'], stopIfTrue=True, font=fn_red, fill=fill_red))

    # Анализ реконструкции ВОЛС
    if rec_df_ is not None:
        df = rec_df_
        rec_df = df
        tz_rec_df = df[df[process_columns['tz_status2']] != 'Исполнена']
        sending_po_rec_df = df[df[process_columns['send_tz_status2']] != 'Исполнена']
        received_po_rec_df = df[df[process_columns['received_tz_status2']] != 'Исполнена']

        ws['G2'] = df[process_columns['plan_date']].count()
        ws['G2'].font = fn_bold
        ws['G2'].alignment = align_center
        ws['G2'].border = border_medium
        # ws['G6'] = sum_sort_month_events(df, process_columns['plan_date'], process_month, last_days_of_month)
        ws['G6'] = rec_df[(rec_df[process_columns['plan_date']] != '') & (rec_df[process_columns['plan_date']] <= last_days_of_month[process_month])][
            process_columns['plan_date']].count()
        ws['G6'].alignment = align_center
        ws['G6'].border = border_medium
        if not args.new_algorithm:
            ws['H6'] = rec_df[(rec_df[process_columns['commissioning_date2']] != '') & (
                    rec_df[process_columns['commissioning_date2']] <= last_days_of_month[process_month]) & (rec_df[process_columns['ks2_date2']] != '') & (
                                      rec_df[process_columns['ks2_date2']] <= last_days_of_month[process_month])][process_columns['commissioning_date2']].count()
        else:
            ws['H6'] = rec_df[(rec_df[process_columns['complete_date2']] != '') & (rec_df[process_columns['complete_date2']] <= last_days_of_month[process_month])][
                process_columns['complete_date2']].count()
        ws['H6'].alignment = align_center
        ws['H6'].border = border_medium

        ws['I6'] = ws['H6'].value - ws['G6'].value
        ws['I6'].alignment = align_center
        ws['I6'].border = border_medium
        ws.conditional_formatting.add('I6', CellIsRule(operator='lessThan', formula=['0'], stopIfTrue=True, font=fn_red, fill=fill_red))
        ws.conditional_formatting.add('I6', CellIsRule(operator='greaterThan', formula=['0'], stopIfTrue=True, font=fn_green, fill=fill_green))
        ws.conditional_formatting.add('I6', CellIsRule(operator='equal', formula=['0'], stopIfTrue=True, font=fn_mag, fill=fill_yellow))

        for i, process in zip(range(10, 19), ['tz_status2',
                                              'send_tz_status2',
                                              'received_tz_status2',
                                              'pir_smr_status2',
                                              'line_scheme_status2',
                                              'tu_status2',
                                              'build_status2',
                                              'ks2_status2',
                                              'commissioning_status2']):
            ws[f'G{i}'] = sum_sort_events(df, process_columns[process], ['Исполнена', 'Не требуется'])
            ws[f'G{i}'].alignment = align_center
            ws[f'G{i}'].border = border_medium
            ws[f'G{i}'].border = border_medium
        for i in range(10, 19):
            ws[f'H{i}'] = ws[f'G{i}'].value - ws['G2'].value
            # ws[f'H{i}'] = f'=G{i}-G2'
            ws[f'H{i}'].alignment = align_center
            ws[f'H{i}'].border = border_medium
            ws.conditional_formatting.add(f'H{i}', CellIsRule(operator='greaterThanOrEqual', formula=['0'], stopIfTrue=True, font=fn_green, fill=fill_green))
            ws.conditional_formatting.add(f'H{i}', CellIsRule(operator='lessThan', formula=['0'], stopIfTrue=True, font=fn_red, fill=fill_red))
    ws = adjust_columns_width(ws)

    # Создание листов для рассылки

    # Создание листа Активные мероприятия строительства месяца отчёта
    # маска для текущего месяца
    if not args.active_year:
        curr_month_bool_mask = (build_dashboard_data[process_columns['plan_date']] <= last_days_of_month[process_month].strftime('%Y-%m-%d'))
    else:
        curr_month_bool_mask = (build_dashboard_data[process_columns['plan_date']] <= last_days_of_month[12].strftime('%Y-%m-%d'))
    # маска для не "Исполнена" или не "Не требуется"
    if not args.new_algorithm:
        curr_status_bool_mask = (~build_dashboard_data[process_columns['commissioning_status']].str.contains('Исполнена|Не требуется', regex=True)) | (
            ~build_dashboard_data[process_columns['ks2_status']].str.contains('Исполнена|Не требуется', regex=True))
    else:
        curr_status_bool_mask = (
                    (build_dashboard_data[process_columns['complete_date']].isna()) & (build_dashboard_data[process_columns['plan_date']] <= last_days_of_month[process_month]))

    # Выборка объектов строительства по маскам
    current_month_build_dataframe = build_dashboard_data[curr_month_bool_mask & curr_status_bool_mask]
    current_month_build_dataframe = current_month_build_dataframe[[process_columns['id'],
                                                                   process_columns['region'],
                                                                   process_columns['name'],
                                                                   process_columns['plan_date'],
                                                                   process_columns['program'],
                                                                   ]]
    current_month_build_dataframe[BP] = BP_BUILD

    # маска для текущего месяца
    if rec_df_ is not None:
        if not args.active_year:
            curr_month_bool_mask = (rec_df[process_columns['plan_date']] <= last_days_of_month[process_month].strftime('%Y-%m-%d'))
        else:
            curr_month_bool_mask = (rec_df[process_columns['plan_date']] <= last_days_of_month[12].strftime('%Y-%m-%d'))
        # маска для не "Исполнена" или не "Не требуется"
        if not args.new_algorithm:
            curr_status_bool_mask = (~rec_df[process_columns['commissioning_status2']].str.contains('Исполнена|Не требуется', regex=True)) | (
                ~rec_df[process_columns['ks2_status2']].str.contains('Исполнена|Не требуется', regex=True))
        else:
            curr_status_bool_mask = (rec_df[process_columns['complete_date2']].isna()) & (rec_df[process_columns['plan_date']] <= last_days_of_month[process_month])

        # Выборка объектов реконструкции по маскам
        current_month_reconstruction_dataframe = rec_df[curr_month_bool_mask & curr_status_bool_mask]
        current_month_reconstruction_dataframe = current_month_reconstruction_dataframe[[process_columns['id'],
                                                                                         process_columns['region'],
                                                                                         process_columns['name'],
                                                                                         process_columns['plan_date'],
                                                                                         process_columns['program'],
                                                                                         ]]
        current_month_reconstruction_dataframe[BP] = BP_RECON  # Добавляем столбец с названием бизнес-процесса

    # Объединяем стройку и реконструкцию
    if rec_df_ is not None:
        current_month_dataframe = pd.concat([current_month_build_dataframe, current_month_reconstruction_dataframe], ignore_index=True).reset_index(drop=True).sort_values(
            by=columns_for_sort)
    else:
        current_month_dataframe = current_month_build_dataframe

    if not current_month_dataframe.empty:
        print(f'Создаем лист отчета: {Color.GREEN}"{report_sheets["current_month"]}"{Color.END}')
        wb.excel_format_table(current_month_dataframe, report_sheets['current_month'], excel_tables_names[report_sheets['current_month']])

    # Создание листа Нет ТЗ
    # формируем таблицы ТЗ для стройки и реконструкции
    #
    tz_build_dataframe = tz_build_dataframe[[process_columns['id'],
                                             process_columns['region'],
                                             process_columns['name'],
                                             process_columns['plan_date'],
                                             process_columns['program'],
                                             ]]
    tz_build_dataframe[BP] = BP_BUILD
    if rec_df_ is not None:
        tz_rec_df = tz_rec_df[[process_columns['id'],
                               process_columns['region'],
                               process_columns['name'],
                               process_columns['plan_date'],
                               process_columns['program'],
                               ]]
        tz_rec_df[BP] = BP_RECON
        # Объединяем ТЗ стройки и реконструкции
        tz_dataframe = pd.concat([tz_build_dataframe, tz_rec_df], ignore_index=True).reset_index(drop=True).sort_values(by=columns_for_sort)
    else:
        tz_dataframe = tz_build_dataframe

    # write_report_table_to_file(tz_dataframe, file_name, report_sheets['tz'], excel_tables_names, excel_cell_names,
    #                            table_style)
    if not tz_dataframe.empty:
        if args.send_email:
            threading.Thread(target=call_send_email, args=(tz_dataframe, reports_data['tz'], args.no_debug, EMAIL_ADDRESS, EMAIL_PASSWORD, date_last_update)).start()
        print(f'Создаем лист отчета: {Color.GREEN}"{report_sheets["tz"]}"{Color.END}')
        wb.excel_format_table(tz_dataframe, report_sheets['tz'], excel_tables_names[report_sheets['tz']])

    # Создание листа Не переданы ТЗ в ПО
    # формируем таблицы передачи в ПО для стройки и реконструкции
    #
    sending_po_build_dataframe = sending_po_build_dataframe[[process_columns['id'],
                                                             process_columns['region'],
                                                             process_columns['name'],
                                                             process_columns['plan_date'],
                                                             process_columns['program'],
                                                             ]]
    sending_po_build_dataframe[BP] = BP_BUILD
    if rec_df_ is not None:
        sending_po_rec_df = sending_po_rec_df[[process_columns['id'],
                                               process_columns['region'],
                                               process_columns['name'],
                                               process_columns['plan_date'],
                                               process_columns['program'],
                                               ]]
        sending_po_rec_df[BP] = BP_RECON
        # Объединяем передачу ТЗ в ПО стройки и реконструкции
        sending_po_dataframe = pd.concat([sending_po_build_dataframe, sending_po_rec_df], ignore_index=True).reset_index(drop=True)
    else:
        sending_po_dataframe = sending_po_build_dataframe
        # Убираем мероприятия с не выданными ТЗ
    sending_po_dataframe = pd.concat([sending_po_dataframe, tz_dataframe], ignore_index=True).drop_duplicates(keep=False).reset_index(drop=True).sort_values(by=columns_for_sort)
    # write_report_table_to_file(sending_po_dataframe, file_name, report_sheets['sending_po'], excel_tables_names,
    #                            excel_cell_names, table_style)
    if not sending_po_dataframe.empty:
        if args.send_email:
            threading.Thread(target=call_send_email,
                             args=(sending_po_dataframe, reports_data['sending_po'], args.no_debug, EMAIL_ADDRESS, EMAIL_PASSWORD, date_last_update)).start()
        print(f'Создаем лист отчета: {Color.GREEN}"{report_sheets["sending_po"]}"{Color.END}')
        wb.excel_format_table(sending_po_dataframe, report_sheets['sending_po'], excel_tables_names[report_sheets['sending_po']])

    # Создание листа ТЗ не принято ПО
    #
    # формируем таблицы не принято ПО для стройки и реконструкции
    #
    received_po_build_dataframe = received_po_build_dataframe[[process_columns['id'],
                                                               process_columns['region'],
                                                               process_columns['name'],
                                                               process_columns['plan_date'],
                                                               process_columns['program'],
                                                               ]]
    received_po_build_dataframe[BP] = BP_BUILD

    if rec_df_ is not None:
        received_po_rec_df = received_po_rec_df[[process_columns['id'],
                                                 process_columns['region'],
                                                 process_columns['name'],
                                                 process_columns['plan_date'],
                                                 process_columns['program'],
                                                 ]]
        received_po_rec_df[BP] = BP_RECON
        # Объединяем не принято в ПО стройки и реконструкции
        received_po_dataframe = pd.concat([received_po_build_dataframe, received_po_rec_df], ignore_index=True).reset_index(drop=True)
    else:
        received_po_dataframe = received_po_build_dataframe
    # Убираем мероприятия с не выданными ТЗ и не переданные в ПО
    received_po_dataframe = pd.concat([received_po_dataframe, sending_po_dataframe, tz_dataframe],
                                      ignore_index=True).drop_duplicates(keep=False).reset_index(drop=True).sort_values(by=columns_for_sort)
    # write_report_table_to_file(received_po_dataframe, file_name, report_sheets['received_po'], excel_tables_names,
    #                            excel_cell_names, table_style)
    if not received_po_dataframe.empty:
        if args.send_email:
            threading.Thread(target=call_send_email,
                             args=(received_po_dataframe, reports_data['received_po'], args.no_debug, EMAIL_ADDRESS, EMAIL_PASSWORD, date_last_update)).start()
        print(f'Создаем лист отчета: {Color.GREEN}"{report_sheets["received_po"]}"{Color.END}')
        wb.excel_format_table(received_po_dataframe, report_sheets['received_po'], excel_tables_names[report_sheets['received_po']])

    if args.soc_report:
        # TODO необходимо сделать подсчет соцсоревнования в соответствии с 2-мя режимами счета на КС-2 и принятию ВОЛС и только по завершению ВОЛС
        #
        # Формируем листы соцсоревнования
        #
        soc_df_build = extended_build_df[[process_columns['region'],
                                          process_columns['plan_date'],
                                          process_columns['complete_date']
                                          ]].copy()
        soc_df_build[BP] = BP_BUILD

        soc_df_rec = rec_df_[[process_columns['region'],
                              process_columns['plan_date'],
                              process_columns['complete_date2']
                              ]].copy().rename(columns=rename_columns)
        soc_df_rec[BP] = BP_RECON

        #
        # Маски для соц соревнования
        #
        mask_soc_plan_build = (soc_df_build[process_columns['plan_date']] <= last_days_of_month[process_month].strftime('%Y-%m-%d'))
        mask_soc_done_build = (soc_df_build[process_columns['complete_date']] <= last_days_of_month[process_month].strftime('%Y-%m-%d'))
        mask_soc_plan_rec = (soc_df_rec[process_columns['plan_date']] <= last_days_of_month[process_month].strftime('%Y-%m-%d'))
        mask_soc_done_rec = (soc_df_rec[process_columns['complete_date']] <= last_days_of_month[process_month].strftime('%Y-%m-%d'))

        # Формируем datasets для соц. соревнования
        soc_df_plan_build = soc_df_build[mask_soc_plan_build]
        soc_df_done_build = soc_df_build[mask_soc_done_build]
        soc_df_plan_rec = soc_df_rec[mask_soc_plan_rec]
        soc_df_done_rec = soc_df_rec[mask_soc_done_rec]

        # Считаем мероприятия плана строительства ВОЛС
        soc_report_plan_build = soc_df_plan_build.groupby([process_columns['region']]).agg(
            {
                process_columns['plan_date']: 'count',
            }
        ).reset_index()
        # Считаем мероприятия факта строительства ВОЛС
        soc_report_done_build = soc_df_done_build.groupby([process_columns['region']]).agg(
            {
                process_columns['complete_date']: 'count',
            }
        ).reset_index()
        soc_report_build = pd.merge(soc_report_plan_build, soc_report_done_build, how='outer', sort=True).fillna(value=0).rename(columns=soc_rename_columns)
        soc_report_build[DELTA_CHAR] = soc_report_build['Факт'] - soc_report_build['План']
        soc_report_build.loc["total"] = soc_report_build.sum(numeric_only=True)
        soc_report_build.at["total", 'Регион/Зона мероприятия'] = "ИТОГО:"
        logger.debug(f'{soc_report_build = }')

        if not soc_report_build.empty:
            print(f'Создаем лист отчета: {Color.GREEN}"{report_sheets["soc_build"]}"{Color.END}')
            wb.excel_format_table(soc_report_build, report_sheets['soc_build'], excel_tables_names[report_sheets['soc_build']])

        # Считаем мероприятия плана Реконструкции ВОЛС
        soc_report_plan_rec = soc_df_plan_rec.groupby([process_columns['region']]).agg(
            {
                process_columns['plan_date']: 'count',
            }
        ).reset_index()
        # Считаем мероприятия факта Реконструкции ВОЛС
        soc_report_done_rec = soc_df_done_rec.groupby([process_columns['region']]).agg(
            {
                process_columns['complete_date']: 'count',
            }
        ).reset_index()
        soc_report_rec = pd.merge(soc_report_plan_rec, soc_report_done_rec, how='outer', sort=True).fillna(value=0).rename(columns=soc_rename_columns)
        soc_report_rec[DELTA_CHAR] = soc_report_rec['Факт'] - soc_report_rec['План']
        soc_report_rec.loc["total"] = soc_report_rec.sum(numeric_only=True)
        soc_report_rec.at['total', 'Регион/Зона мероприятия'] = 'ИТОГО:'
        logger.debug(f'{soc_report_rec = }')

        if not soc_report_rec.empty:
            print(f'Создаем лист отчета: {Color.GREEN}"{report_sheets["soc_rec"]}"{Color.END}')
            wb.excel_format_table(soc_report_rec, report_sheets['soc_rec'], excel_tables_names[report_sheets['soc_rec']])
            logger.debug(f'{wb.ws["B2"].value = }')
    #
    # Записываем сформированный файл отчета
    #
    logger.info(f'Удаляем лист {ws_first}')
    wb.remove(ws_first)
    if Path(file_name).is_file():
        try:
            print(f'Удаляем существующий файл отчета {Color.GREEN}"{file_name}"{Color.END}')
            os.remove(file_name)
        except Exception as ex:
            logger.error(f'Ошибка удаления файла: {ex}')
            sys.exit(1)

    try:
        print(f'Сохраняем отформатированные данные в файл {Color.GREEN}"{file_name}"{Color.END}')
        wb.save(file_name)
    except Exception as ex:
        logger.error(f'Ошибка сохранения файла файла: {ex}')
        sys.exit(2)


if __name__ == '__main__':
    main()
